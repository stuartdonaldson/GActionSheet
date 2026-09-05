"""
test_archive.py — GTaskSheet-d33z

Journey-based archive lifecycle: a real action item created in a doc, synced
to the sheet (acquiring globalId + fileId), closed, aged, then swept by
ArchiveManager. Verifies the archived row retains globalId and fileId intact.

The archive sweep also acts as the session teardown — after archival the row
is gone from Actions and the journey doc is effectively retired.
"""
import io
import pytest
import openpyxl

from scn.ai import ai
from scn.engine import CheckpointKind, Surface
from scn.session import ScenarioSession
from tests.helpers.download import download_xlsx

SHEET = Surface.SHEET
STEP = CheckpointKind.STEP

_ARCHIVE_ACTION_TEXT = "d33z archive lifecycle action"


def _find_archive_row(sheet_id: str, action_text: str) -> dict | None:
    """Return the first Archive tab row whose Action column matches action_text, or None."""
    xlsx = download_xlsx(sheet_id)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    if "Archive" not in wb.sheetnames:
        return None
    ws = wb["Archive"]
    col_names = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < len(col_names):
            continue
        row_dict = dict(zip(col_names, row))
        if row_dict.get("Action") == action_text:
            return row_dict
    return None


@pytest.fixture
def scn(settings, request):
    # gts-u6ew.12 (F7): request=request wires JUnit ac.*/ep.* emission to this
    # test node (T24) — without it the session is a no-op reporter and no
    # coverage properties reach pytest.xml (q37d, see test_menu_entry_points.py).
    # Function-scoped (was module-scoped): this file has exactly one test, so
    # scope is behaviorally identical, and a module-scoped `request.node` is a
    # Module collector, not the Function item JUnit properties attach to.
    s = ScenarioSession.new_doc(settings, request=request)
    yield s
    s.close()


def test_archive_lifecycle(scn, settings):
    """GTaskSheet-d33z: row created via full lifecycle is archived with globalId + fileId."""
    sheet_id = settings["testSheetId"]

    # SETUP: append action, sync → row lands in sheet with proper globalId + fileId
    target = ai(action=_ARCHIVE_ACTION_TEXT)
    scn.append_paragraph(target.as_text())
    scn.sync()

    # Pin the globalId so we can address the row by it
    rows = scn.find_sheet_actions()
    matching = [r for r in rows if r.action == _ARCHIVE_ACTION_TEXT]
    assert len(matching) == 1, (
        f"[d33z] Expected 1 sheet row matching action text after sync, got {len(matching)}"
    )
    target.action_id = matching[0].action_id

    # ACT: close the action and age the modified_date so it qualifies for archival
    scn.edit_sheet(target, status="Closed")
    scn.sync()   # Dirty → sheet-wins → doc updated; row is now Closed in sheet

    global_id = f"{scn.doc_id}/{target.action_id}"
    scn._post_fixture("backdate_action_row", {"globalId": global_id, "daysAgo": 35})

    # Run the archive sweep
    scn._post_fixture("archive_journey")

    # ASSERT 1: row absent from Actions
    actions_after = scn.sheet_rows()
    assert not any(r.action == _ARCHIVE_ACTION_TEXT for r in actions_after), (
        f"[d33z] '{_ARCHIVE_ACTION_TEXT}' row still present in Actions after archive sweep"
    )

    # ASSERT 2: row present in Archive with globalId and fileId populated
    archived = _find_archive_row(sheet_id, _ARCHIVE_ACTION_TEXT)
    assert archived is not None, (
        f"[d33z] '{_ARCHIVE_ACTION_TEXT}' row not found in Archive tab after sweep"
    )
    assert archived.get("globalId"), (
        f"[d33z] archived row missing globalId; got {archived.get('globalId')!r}"
    )
    assert archived.get("File Id"), (
        f"[d33z] archived row missing File Id; got {archived.get('File Id')!r}"
    )

    # gts-u6ew.12 (F7): route the full-lifecycle result through
    # expect_callable/checkpoint so tag= reaches the T24 report — same checks
    # already made above, no new assertion — scn/contract.AC_REGISTRY
    # "archive lifecycle".
    def _archive_lifecycle_held() -> str | None:
        if any(r.action == _ARCHIVE_ACTION_TEXT for r in scn.sheet_rows()):
            return f"[d33z] '{_ARCHIVE_ACTION_TEXT}' row still present in Actions after archive sweep"
        row = _find_archive_row(sheet_id, _ARCHIVE_ACTION_TEXT)
        if row is None:
            return f"[d33z] '{_ARCHIVE_ACTION_TEXT}' row not found in Archive tab after sweep"
        if not row.get("globalId"):
            return f"[d33z] archived row missing globalId; got {row.get('globalId')!r}"
        if not row.get("File Id"):
            return f"[d33z] archived row missing File Id; got {row.get('File Id')!r}"
        return None

    scn.expect_callable(
        _archive_lifecycle_held, on=SHEET, tag="archive lifecycle",
    )
    scn.checkpoint(STEP)
