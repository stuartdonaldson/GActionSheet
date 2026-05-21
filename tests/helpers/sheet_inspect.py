"""openpyxl-based helpers for asserting tracking sheet content from an xlsx download."""
import io
import datetime
import openpyxl


def load_sheet(xlsx_bytes: bytes, sheet_name: str = None):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return wb[sheet_name] if sheet_name else wb.active


def headers(ws) -> dict[str, int]:
    """Return {header_name: col_index} from row 1 (1-based)."""
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def rows_as_dicts(ws) -> list[dict]:
    """Return all data rows (skipping header) as dicts keyed by header name."""
    cols = headers(ws)
    result = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(c.value is None for c in row):
            continue
        result.append({name: row[idx - 1].value for name, idx in cols.items()})
    return result


def find_row(ws, doc_url: str, action_id: int) -> dict | None:
    """Find a sheet row by document URL (hyperlink match) and ID.

    Handles both proper Excel hyperlink objects and =HYPERLINK() formula strings,
    since Google Sheets exports formula-based hyperlinks without an XML hyperlink
    element that openpyxl can read via cell.hyperlink.
    """
    import re as _re
    _HYPERLINK_FORMULA_RE = _re.compile(r'^=HYPERLINK\("([^"]+)"', _re.IGNORECASE)
    cols = headers(ws)
    id_col = cols.get("ID")
    doc_col = cols.get("Document")
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_id = row[id_col - 1].value
        cell_doc = row[doc_col - 1]
        # Resolve URL from XML hyperlink attribute OR =HYPERLINK() formula string.
        if cell_doc.hyperlink:
            url = cell_doc.hyperlink.target
        elif isinstance(cell_doc.value, str):
            m = _HYPERLINK_FORMULA_RE.match(cell_doc.value)
            url = m.group(1) if m else cell_doc.value
        else:
            url = None
        if cell_id == action_id and url and doc_url in url:
            return {name: row[idx - 1].value for name, idx in cols.items()}
    return None


def assert_date_cell(ws, row_dict: dict, col_name: str):
    """Assert that a named cell in the matched row is stored as a native Date value."""
    cols = headers(ws)
    col_idx = cols[col_name]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[cols["ID"] - 1].value == row_dict["ID"]:
            cell = row[col_idx - 1]
            assert isinstance(cell.value, datetime.datetime), (
                f"{col_name} cell is {type(cell.value).__name__}, expected datetime"
            )
            return
    raise AssertionError(f"Row with ID={row_dict['ID']} not found")
