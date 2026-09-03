"""
surfaces.py — DOC, SHEET, TRACKER surface readers (GTaskSheet-5vwu.6).

Spec: docs/atdd/atdd-lifecycle.md §16.5, §16.3 rule 4
Design: docs/atdd/scenario-harness-design.md §3.7

Each reader returns plain ai-shaped records; no assertion logic lives here.
Column mapping comes from scn/contract.py — never restated in this module.
"""
import io
import re
from typing import TYPE_CHECKING

import docx as _docx_pkg
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
import openpyxl

from scn.ai import ai
from scn import contract as _contract

_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_EMAIL_RE = re.compile(r"[\w.+\-]+@[\w\-]+(?:\.[a-z]{2,})+", re.IGNORECASE)
_HYPERLINK_FORMULA_RE = re.compile(r'^=HYPERLINK\("([^"]+)"(?:,"([^"]*)")?\)', re.IGNORECASE)
_GDOC_ID_RE = re.compile(r"/document/d/([^/]+)/")
_DRIVE_ID_RE = re.compile(r"[?&]id=([a-zA-Z0-9_-]+)")
_TRACKER_HEADING = "Action Item Summary"
_TRACKER_HEADING_OLD = "=== Tracked Actions ==="  # legacy; accepted for back-compat


def _iter_block_items(document):
    """Yield paragraphs and tables from the document body in order."""
    body = document.element.body
    for child in body:
        if child.tag == qn("w:p"):
            yield Paragraph(child, document)
        elif child.tag == qn("w:tbl"):
            yield Table(child, document)


def _cell_email_by_rid(cell) -> dict[str, str]:
    """Build rId → email map from a table cell's hyperlink relationships (via cell part)."""
    import urllib.parse as _urlparse
    result: dict[str, str] = {}
    try:
        part = cell.part
        for rel in part.rels.values():
            if "hyperlink" not in rel.reltype:
                continue
            url = rel.target_ref or ""
            m = re.match(r"mailto:([^\?&]+)", url, re.IGNORECASE)
            if m:
                result[rel.rId] = m.group(1).strip()
                continue
            m = re.search(r"[?&]email=([^&]+)", url, re.IGNORECASE)
            if m:
                result[rel.rId] = _urlparse.unquote(m.group(1)).strip()
    except Exception:
        pass
    return result


def _find_chip_in_cell(cell) -> str | None:
    """Return the chip email from the first hyperlink with an email URL in a table cell, or None."""
    email_map = _cell_email_by_rid(cell)
    for para in cell.paragraphs:
        for hl in para._element.findall(qn("w:hyperlink")):
            r_id = hl.get(f"{{{_R_NS}}}id", "")
            if r_id in email_map:
                return email_map[r_id]
    return None


def _find_action_url_in_cell(cell) -> str | None:
    """Return the first non-email hyperlink URL from a table cell, or None."""
    try:
        part = cell.part
        for rel in part.rels.values():
            if "hyperlink" not in rel.reltype:
                continue
            url = rel.target_ref or ""
            if not url.startswith("mailto:") and "email=" not in url:
                return url
    except Exception:
        pass
    return None


_ASSIGNEE_SOURCE_MAP = {"chip": "chip", "text": "parsed", None: None}


class DocReader:
    """Read floating-action ai records from a .docx (DOC surface, §16.5).

    Derives its records from ``tests.helpers.doc_inspect.floating_actions`` —
    the full ADR-0027 grammar parser (gts-1ej4) — rather than maintaining a
    second, weaker parser here. Detection is still the ACT-N:/AI-N: text
    token (ADR-0008; dual-prefix per ADR-0023 — ACT-N: canonical on write,
    AI-N: permanently read-compatible). Returns all matching paragraphs
    including multiple occurrences of the same action_id — the
    engine/assertions enforce the identical-occurrence invariant. A bare
    ``AI:``/``ACT:`` trigger (no number yet) and an unparseable paragraph
    (ADR-0027 rule 6) are not "established" actions and are excluded, matching
    this reader's pre-convergence behavior.
    """

    def read(self, docx_bytes: bytes, doc_id: str) -> list[ai]:
        from tests.helpers import doc_inspect

        document = doc_inspect.load_doc(docx_bytes)
        results = []
        for parsed in doc_inspect.floating_actions(document):
            if parsed.token is None:  # bare trigger or unparseable (rule 6)
                continue
            results.append(ai(
                action=parsed.action_text,
                assignee=parsed.assignee_email,
                action_id=parsed.token,
                status=parsed.status,
                assignee_source=_ASSIGNEE_SOURCE_MAP[parsed.assignee_source],
            ))
        return results


class SheetReader:
    """Read action ai records from a .xlsx ActionSheet (SHEET surface, §16.5).

    Scoped to doc_id (§16.3 rule 4). Column mapping from scn/contract.
    doc_id and doc_name are DERIVED from document_formula (col 7), not stored
    columns (Coordination Log §.1 finding #1).
    """

    def read(self, xlsx_bytes: bytes, doc_id: str, tab_name: str = "Actions") -> list[ai]:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[tab_name] if tab_name in wb.sheetnames else wb.active

        # Build 1-based column index from contract (authoritative)
        col = _contract.COLUMNS_BY_FIELD  # e.g. {"global_id": 1, "action_id": 2, ...}
        doc_col = col["document_formula"]  # column 7

        results = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if all(c.value is None for c in row):
                continue

            # Derive doc_id and doc_name from the document_formula cell
            doc_cell = row[doc_col - 1]
            derived_doc_id, derived_doc_name = self._parse_document_formula(doc_cell)

            if derived_doc_id != doc_id:
                continue

            def _cell_val(field: str):
                idx = col.get(field)
                if idx is None or idx > len(row):
                    return None  # short row: a trailing optional column (custom_fields)
                c = row[idx - 1]
                v = c.value
                return str(v) if v is not None else None

            obj = ai(
                action=_cell_val("action_text") or "",
                assignee=_cell_val("assignee_email") or None,
                action_id=_cell_val("action_id") or None,
                status=_cell_val("status") or None,
                assignee_source=None,
            )
            obj.global_id = _cell_val("global_id")
            obj.assignee_name = _cell_val("assignee_name")
            obj.sync_status = _cell_val("sync_status") or ""
            # ADR-0027 rule 9: additive, optional column; raw JSON text here, decoded
            # by the consumer (tests/helpers/doc_sheet_agreement.py) rather than in
            # this reader, which stays a plain cell-to-record mapper.
            obj.custom_fields = _cell_val("custom_fields")
            obj.doc_id = derived_doc_id
            obj.doc_name = derived_doc_name
            # created_date/modified_date: keep the raw cell value (a naive datetime,
            # spreadsheet-local wall clock — Sheets/xlsx datetimes carry no zone info)
            # rather than stringifying, so assertions._normalize_date can convert it
            # for comparison against the ISO-8601 UTC strings the webapp returns.
            obj.created_date = row[col["created_date"] - 1].value
            obj.modified_date = row[col["modified_date"] - 1].value
            results.append(obj)

        return results

    @staticmethod
    def _parse_document_formula(cell) -> tuple[str | None, str | None]:
        """Extract (doc_id, doc_name) from a Document column cell.

        Handles both openpyxl hyperlink objects and =HYPERLINK() formula strings.
        """
        doc_id = None
        doc_name = None

        # Path 1: openpyxl-resolved hyperlink object
        if cell.hyperlink:
            url = cell.hyperlink.target or ""
            m = _GDOC_ID_RE.search(url) or _DRIVE_ID_RE.search(url)
            if m:
                doc_id = m.group(1)
            doc_name = cell.hyperlink.tooltip or (str(cell.value) if cell.value else None)

        # Path 2: =HYPERLINK("url","name") formula string.
        # Google Sheets normalises document URLs to https://docs.google.com/open?id=DOCID
        # when storing =HYPERLINK() formulas, so both URL formats must be matched.
        if doc_id is None and isinstance(cell.value, str):
            fm = _HYPERLINK_FORMULA_RE.match(cell.value)
            if fm:
                url = fm.group(1)
                m = _GDOC_ID_RE.search(url) or _DRIVE_ID_RE.search(url)
                if m:
                    doc_id = m.group(1)
                if doc_name is None:
                    doc_name = fm.group(2)  # may be None if no name arg

        return doc_id, doc_name


class TrackerReader:
    """Read action ai records from the tracker table in a .docx (TRACKER surface, §16.5).

    The table lives after a paragraph reading '=== Tracked Actions ==='.
    The assignee is rendered as a person chip (hyperlink) in the 'Assignee Name'
    column in production-synced docs; plain text in fixture-built docs.
    """

    def read(self, docx_bytes: bytes, doc_id: str) -> list[ai]:
        document = _docx_pkg.Document(io.BytesIO(docx_bytes))
        return self._parse_tracker(document)

    def _parse_tracker(self, document) -> list[ai]:
        found_heading = False
        for block in _iter_block_items(document):
            if isinstance(block, Paragraph):
                if block.text.strip() in (_TRACKER_HEADING, _TRACKER_HEADING_OLD):
                    found_heading = True
                    continue
            elif isinstance(block, Table) and found_heading:
                return self._parse_table(block)
        return []

    def _parse_table(self, table) -> list[ai]:
        if not table.rows:
            return []
        headers = [cell.text.strip() for cell in table.rows[0].cells]
        col_idx: dict[str, int] = {h: i for i, h in enumerate(headers)}

        results = []
        for row in table.rows[1:]:
            cells = row.cells
            if all(c.text.strip() == "" for c in cells):
                continue

            action_id = _col_text(cells, col_idx, "ID") or None
            action_text = _col_text(cells, col_idx, "Action") or ""
            status = _col_text(cells, col_idx, "Status") or None
            id_url = None
            if "ID" in col_idx:
                id_url = _find_action_url_in_cell(cells[col_idx["ID"]])

            # Assignee: prefer chip in 'Assignee Name' (legacy schema) or 'Assignee'
            # (current GAS schema = ['ID','Assignee','Action','Status']).
            chip_email = None
            assignee_name_text = ""
            assignee_col = "Assignee Name" if "Assignee Name" in col_idx else "Assignee"
            if assignee_col in col_idx:
                name_cell = cells[col_idx[assignee_col]]
                chip_email = _find_chip_in_cell(name_cell)
                assignee_name_text = name_cell.text.strip()

            if chip_email:
                assignee = chip_email
                assignee_source = "chip"
            else:
                # fall back to 'Assignee Email' column (legacy schema) or plain cell text
                assignee = _col_text(cells, col_idx, "Assignee Email") or None
                if not assignee and assignee_name_text and _EMAIL_RE.fullmatch(assignee_name_text):
                    assignee = assignee_name_text
                assignee_source = "parsed" if assignee else None

            obj = ai(
                action=action_text,
                assignee=assignee,
                action_id=action_id,
                status=status,
                assignee_source=assignee_source,
            )
            obj.assignee_name = assignee_name_text or _col_text(cells, col_idx, "Assignee Name")
            obj.id_url = id_url
            results.append(obj)

        return results


def _col_text(cells: list, col_idx: dict[str, int], header: str) -> str | None:
    idx = col_idx.get(header)
    if idx is None or idx >= len(cells):
        return None
    return cells[idx].text.strip() or None
